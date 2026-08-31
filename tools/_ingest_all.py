import os, re, json, shutil
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
SOLV_DIR = os.path.join(BASE, "..")
DATA_JS = os.path.join(SOLV_DIR, "data.js")
DATA_DIR = os.path.join(SOLV_DIR, "data")
HEADER_COMMENT = "// 保险行业偿付能力分析平台 - 数据文件（自动生成，勿手改；更新请运行 data_pipeline.py）"

# ============================================================
# 工具
# ============================================================
def normalize_period(s):
    s = str(s).strip()
    m = re.match(r'(\d{4})[^第]*第?([1-4])季度', s)
    if m: return f"{m.group(1)}Q{m.group(2)}"
    m = re.match(r'(\d{4})', s)
    return m.group(1) if m else s

def make_period_def(key):
    m = re.match(r'(\d{4})Q([1-4])', key)
    if m:
        y, q = int(m.group(1)), int(m.group(2))
        return {"key": key, "year": y, "q": q, "kind": "quarter", "label": f"{y}年第{q}季度", "source": "quarter"}
    m = re.match(r'(\d{4})$', key)
    if m:
        y = int(m.group(1))
        return {"key": key, "year": y, "q": 4, "kind": "year-end", "label": f"{y}年", "source": "prelim"}
    return {"key": key, "year": 0, "q": 0, "kind": "other", "label": key, "source": "other"}

# ============================================================
# 读取偿付能力主表 (C..V)
# ============================================================
SOLV_COL_MAP = {
    "综合偿付能力充足率": "C", "核心偿付能力充足率": "D",
    "综合偿付能力溢额": "E", "核心偿付能力溢额": "F",
    "认可资产": "G", "认可负债": "H", "实际资本": "I",
    "核心一级资本": "J", "核心二级资本": "K", "附属一级资本": "L", "附属二级资本": "M",
    "最低资本": "N", "量化风险最低资本": "O", "寿险风险最低资本": "P",
    "非寿险风险最低资本": "Q", "市场风险最低资本": "R", "信用风险最低资本": "S",
    "量化风险分散效应": "T", "特定类别保险合同损失吸收效应": "U", "控制风险最低资本": "V",
}

def read_solvency(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = None
    for ri, r in enumerate(rows):
        if r and any(str(c).strip() in SOLV_COL_MAP for c in r if c):
            hdr_i = ri; break
    if hdr_i is None:
        print(f"  [WARN] 未找到表头: {os.path.basename(path)}"); wb.close(); return {}
    headers = [str(c).strip() if c else "" for c in rows[hdr_i]]
    col_idx = {SOLV_COL_MAP[h]: ci for ci, h in enumerate(headers) if h in SOLV_COL_MAP}
    result = {}
    for ri in range(hdr_i + 1, len(rows)):
        r = rows[ri]
        if not r or not r[0]: continue
        company = str(r[0]).strip()
        if len(r) < 2 or not r[1]: continue
        pk = normalize_period(r[1])
        vals = {}
        for f, ci in col_idx.items():
            if ci < len(r) and r[ci] is not None:
                try: vals[f] = float(r[ci])
                except: pass
        if vals:
            result.setdefault(company, {})[pk] = vals
    wb.close()
    return result

# ============================================================
# 读取最低资本子风险 (mc*)
# ============================================================
MC_METRIC_MAP = {
    "寿险业务保险风险-损失发生风险最低资本": "mcP_loss",
    "寿险业务保险风险-退保风险最低资本": "mcP_surr",
    "寿险业务保险风险-费用风险最低资本": "mcP_exp",
    "寿险业务保险风险-风险分散效应": "mcP_div",
    "非寿险业务保险风险-保费及准备金风险最低资本": "mcQ_prem",
    "非寿险业务保险风险-巨灾风险最低资本": "mcQ_cata",
    "非寿险业务保险风险-风险分散效应": "mcQ_div",
    "市场风险-利率风险最低资本": "mcR_rate",
    "市场风险-权益价格风险最低资本": "mcR_eq",
    "市场风险-房地产价格风险最低资本": "mcR_re",
    "市场风险-境外固定收益类资产价格风险最低资本": "mcR_ofb",
    "市场风险-境外权益类资产价格风险最低资本": "mcR_ofe",
    "市场风险-汇率风险最低资本": "mcR_fx",
    "市场风险-风险分散效应": "mcR_div",
    "信用风险-利差风险最低资本": "mcS_spr",
    "信用风险-交易对手违约风险最低资本": "mcS_def",
    "信用风险-风险分散效应": "mcS_div",
    "损失吸收调整-不考虑上限": "mcU_nocap",
    "损失吸收效应调整上限": "mcU_cap",
    "附加资本": "mcAdd",
    "逆周期附加资本": "mcAdd_cyc",
    "D-SII附加资本": "mcAdd_dsii",
    "G-SII附加资本": "mcAdd_gsii",
    "其他附加资本": "mcAdd_oth",
}

def read_mincap(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        print(f"  [WARN] 行数不足: {os.path.basename(path)}"); wb.close(); return {}
    # 列→公司 (row0 标记块起始)
    col_company = {}
    cur = None
    for ci, c in enumerate(rows[0]):
        if c and str(c).strip():
            cur = str(c).strip()
        col_company[ci] = cur
    # 列→期次 (row1)
    col_period = {}
    for ci, p in enumerate(rows[1]):
        if p:
            col_period[ci] = normalize_period(str(p).strip())
    result = {}
    for ri in range(2, len(rows)):
        r = rows[ri]
        if not r or not r[0]: continue
        field = MC_METRIC_MAP.get(str(r[0]).strip())
        if not field: continue
        for ci in range(1, len(r)):
            company = col_company.get(ci)
            period = col_period.get(ci)
            if not company or not period: continue
            if r[ci] is None: continue
            try: val = float(r[ci])
            except: continue
            result.setdefault(company, {}).setdefault(period, {})[field] = val
    wb.close()
    return result

# ============================================================
# 主流程
# ============================================================
SOLV_FILES = [
    ("集团偿付能力.xlsx", "group"),
    ("人身险偿付能力.xlsx", "life"),
    ("产险偿付能力.xlsx", "property"),
    ("再保偿付能力.xlsx", "reins"),
]
MC_FILES = [
    ("人身险最低资本.xlsx", "life"),
    ("产险最低资本.xlsx", "property"),
    ("再保最低资本.xlsx", "reins"),
]

# 读取现有 data.js
txt = open(DATA_JS, encoding="utf-8").read()
m = re.search(r"let SOLVENCY_DATA\s*=\s*(\{.*\})\s*;", txt, re.S)
D = json.loads(m.group(1))

# 读取源
src_main = {}
src_mc = {}
print("=== 读取偿付能力主表 ===")
for fname, seg in SOLV_FILES:
    p = os.path.join(DATA_DIR, fname)
    if not os.path.exists(p): print(f"[SKIP] {fname}"); continue
    print(f"  {fname} ({seg}) ...", end="")
    d = read_solvency(p)
    n_comp = len(d)
    n_periods = sum(len(v) for v in d.values())
    print(f" {n_comp}公司/{n_periods}期")
    for c, pers in d.items():
        src_main.setdefault(c, {})
        for pk, v in pers.items():
            src_main[c].setdefault(seg, {})[pk] = v

print("=== 读取最低资本子风险 ===")
for fname, seg in MC_FILES:
    p = os.path.join(DATA_DIR, fname)
    if not os.path.exists(p): print(f"[SKIP] {fname}"); continue
    print(f"  {fname} ({seg}) ...", end="")
    d = read_mincap(p)
    n_comp = len(d)
    n_periods = sum(len(v) for v in d.values())
    print(f" {n_comp}公司/{n_periods}期")
    for c, pers in d.items():
        src_mc.setdefault(c, {})
        for pk, v in pers.items():
            src_mc[c].setdefault(seg, {})[pk] = v

# 统计变更
stats = {"companies_added": 0, "periods_added": 0, "periods_replaced": 0, "companies_new": 0}

# 合并主数据
for seg, obj in D["segments"].items():
    existing_companies = set(obj["companies"])
    for c, segdata in src_main.items():
        if seg not in segdata: continue
        # 新公司
        if c not in existing_companies:
            obj["companies"].append(c)
            obj["data"][c] = {}
            stats["companies_added"] += 1
            existing_companies.add(c)
        if c not in obj["data"]:
            obj["data"][c] = {}
        for pk, vals in segdata[seg].items():
            if pk in obj["data"][c] and obj["data"][c][pk]:
                stats["periods_replaced"] += 1
            else:
                stats["periods_added"] += 1
            obj["data"][c][pk] = vals
    # 合并 MC
    for c, segdata in src_mc.items():
        if seg not in segdata: continue
        if c not in obj["data"]:
            obj["data"][c] = {}
        for pk, vals in segdata[seg].items():
            if pk not in obj["data"][c] or not obj["data"][c][pk]:
                obj["data"][c][pk] = {}
            obj["data"][c][pk].update(vals)
    # 重建 periods 数组
    all_keys = set()
    for c, pdata in obj["data"].items():
        for pk in pdata:
            if isinstance(pdata[pk], dict) and pdata[pk]:
                all_keys.add(pk)
    old_defs = {p["key"]: p for p in obj["periods"]}
    new_periods = []
    for k in sorted(all_keys):
        if k in old_defs:
            new_periods.append(old_defs[k])
        else:
            new_periods.append(make_period_def(k))
            stats["periods_added"] += 1
    new_periods.sort(key=lambda x: (x["year"], x["q"]))
    obj["periods"] = new_periods

# 写回（保留头部注释）
out = HEADER_COMMENT + "\nlet SOLVENCY_DATA = " + json.dumps(D, ensure_ascii=False, separators=(",", ":")) + ";\n"
import os as _os
if _os.environ.get("DRY"):
    print("\n[DRY-RUN] 未写入文件")
else:
    shutil.copy2(DATA_JS, DATA_JS + ".bak")
    with open(DATA_JS, "w", encoding="utf-8") as f:
        f.write(out)

print("\n=== 合并统计 ===")
print(f"新增公司: {stats['companies_added']}")
print(f"新增期次(含periods数组): {stats['periods_added']}")
print(f"替换期次(历史): {stats['periods_replaced']}")
print(f"data.js 大小: {len(out)/1024/1024:.2f} MB")
print("备份: data.js.bak")
