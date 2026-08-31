# -*- coding: utf-8 -*-
"""
单季度数据更新一键脚本（覆盖全数据产物，替代手工逐项编辑）。

设计要点：
  * 【自动模式（默认）】扫描 data/ 目录下所有 .xlsx/.xls，从文件名自动识别
    「板块类型 + 期次 + 修订号」，批量更新所有能识别的文件；同一板块同期次
    若有多个修订版，只取修订号最大的一份。无期次标记的旧历史文件自动跳过，
    绝不误伤。这样每次季度更新只需两步：把新 Excel 拖进 data/ → 跑一条命令。
  * 手动模式：--rev / --quarter 指定修订号与期次，只处理文件名含 rev 的文件
    （老行为，兼容）。
  * 默认 --overwrite：用源文件里该季度的值【覆盖】目标公司在 data 中的对应期次记录
    （仅覆盖该季度、仅覆盖源文件中出现的公司，绝不触碰其它期次/其它公司）。
  * 加法安全：即便带 --overwrite，也不会删除任何历史期次或无关公司。
  * 最低资本明细(MC)：从“人身险/产险/再保 最低资本{rev}.xlsx”抽取 P~V+N 合并进 data.js
    （仅覆盖这几个字段，其余 C..V 不动；寿险 MC 在偿付能力主表之后写入，作为权威来源）。

覆盖产物：data.js / reg_industry.js / life_capital_detail.js / property_capital_detail.js

用法：
  python tools/update_quarter.py                       # 自动模式：扫描 data/ 全部新文件并更新
  python tools/update_quarter.py --dry                # 只打印将处理的文件与条数，不写文件
  python tools/update_quarter.py --rev 0831           # 手动模式：只处理文件名含 0831 的文件
  python tools/update_quarter.py --quarter 2026Q2     # 手动模式：指定目标期次
  python tools/update_quarter.py --deploy             # 更新 + 同步 dist + 构建单文件 + git 推送 + curl 校验
  python tools/update_quarter.py --no-overwrite       # 仅填补缺失（已有则跳过），不覆盖
"""
import argparse, json, os, re, shutil, subprocess, sys, glob
import json5
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SOLV = os.path.dirname(HERE)
DATA = os.path.join(SOLV, "data")
DATA_JS = os.path.join(SOLV, "data.js")
REG_JS = os.path.join(SOLV, "reg_industry.js")
LIFE_CAP_JS = os.path.join(SOLV, "life_capital_detail.js")
PROP_CAP_JS = os.path.join(SOLV, "property_capital_detail.js")

SOLV_COL_MAP = {
    "综合偿付能力充足率": "C", "核心偿付能力充足率": "D",
    "综合偿付能力溢额": "E", "核心偿付能力溢额": "F",
    "认可资产": "G", "认可负债": "H", "实际资本": "I",
    "核心一级资本": "J", "核心二级资本": "K", "附属一级资本": "L", "附属二级资本": "M",
    "最低资本": "N", "量化风险最低资本": "O", "寿险风险最低资本": "P",
    "非寿险风险最低资本": "Q", "市场风险最低资本": "R", "信用风险最低资本": "S",
    "量化风险分散效应": "T", "特定类别保险合同损失吸收效应": "U", "控制风险最低资本": "V",
}
CAP_ROW_MAP = {
    "核心一级资本": "core1", "净资产": "net", "对净资产的调整额": "adj",
    "各项非认可资产的账面价值": "adj_nonrec",
    "长期股权投资的认可价值与账面价值的差额": "adj_lti",
    "投资性房地产（包括保险公司以物权方式或通过子公司等方式持有的投资性房地产）的公允价值增值（扣除减值、折旧及所得税影响）": "adj_invprop",
    "递延所得税资产（由经营性亏损引起的递延所得税资产除外）": "adj_dta",
    "对农业保险提取的大灾风险准备金": "adj_cat",
    "计入核心一级资本的保单未来盈余": "adj_fvs",
    "符合核心一级资本标准的负债类资本工具且按规定可计入核心一级资本的金额": "adj_liab",
    "银保监会规定的其他调整项目": "adj_other",
    "核心二级资本": "core2", "优先股": "c2_pref",
    "计入核心二级资本的保单未来盈余": "c2_fvm", "其他核心二级资本": "c2_oth",
    "减：超限额应扣除的部分": "c2_ded",
    "附属一级资本": "sub1", "附属二级资本": "sub2", "实际资本合计": "total",
}
REG_SEG_MAP = {"财产险公司": "property", "人身险公司": "life", "再保险公司": "reins"}
# 最低资本表：聚合项（与 data.js 主表 P~V+N 对应）按指标名匹配
MC_AGG_MAP = {
    "寿险业务保险风险最低资本合计": "P",
    "非寿险业务保险风险最低资本合计": "Q",
    "市场风险-最低资本合计": "R",
    "信用风险-最低资本合计": "S",
    "量化风险分散效应": "T",
    "特定类别保险合同损失吸收效应": "U",
    "控制风险最低资本": "V",
    "最低资本": "N",
}
# 最低资本表：下钻明细项（写入 segments.<seg>.mcDetail）按指标名匹配
MC_DETAIL_MAP = {
    "寿险业务保险风险-损失发生风险最低资本": "mcP_loss",
    "寿险业务保险风险-退保风险最低资本": "mcP_surr",
    "寿险业务保险风险-费用风险最低资本": "mcP_exp",
    "寿险业务保险风险-风险分散效应": "mcP_div",
    "非寿险业务保险风险-保费及准备金风险最低资本": "mcQ_prem",
    "非寿险业务保险风险-巨灾风险最低资本": "mcQ_cata",
    "非寿险业务保险风险-风险分散效应": "mcQ_div",
    "市场风险-利率风险最低资本": "mcR_rate",
    "市场风险-权益价格风险最低资本": "mcR_eq",
    "市场风险-房地产价格风险最低资本": "mcR_re",
    "市场风险-境外固定收益类资产价格风险最低资本": "mcR_ofb",
    "市场风险-境外权益类资产价格风险最低资本": "mcR_ofe",
    "市场风险-汇率风险最低资本": "mcR_fx",
    "市场风险-风险分散效应": "mcR_div",
    "信用风险-利差风险最低资本": "mcS_spr",
    "信用风险-交易对手违约风险最低资本": "mcS_def",
    "信用风险-风险分散效应": "mcS_div",
    "附加资本": "mcAdd",
    "逆周期附加资本": "mcAdd_cyc",
    "D-SII附加资本": "mcAdd_dsii",
    "G-SII附加资本": "mcAdd_gsii",
    "其他附加资本": "mcAdd_oth",
}
# 旧版最低资本表兜底行号（无指标名匹配时回退）
MC_ROW_MAP = {3: "P", 8: "Q", 12: "R", 20: "S", 24: "T", 25: "U", 28: "V", 34: "N"}


def period_def(key):
    m = re.match(r"(\d{4})Q([1-4])$", key)
    if m:
        y, q = int(m.group(1)), int(m.group(2))
        return {"key": key, "year": y, "q": q, "kind": "quarter",
                "label": f"{y}年第{q}季度", "source": "quarter"}
    m = re.match(r"(\d{4})$", key)
    if m:
        y = int(m.group(1))
        return {"key": key, "year": y, "q": 4, "kind": "year-end",
                "label": f"{y}年", "source": "prelim"}
    raise SystemExit(f"[ERR] 无法解析期次键: {key}")


def rev_files(base, rev):
    """在 data/ 下找 base*.xlsx，只返回文件名含 rev 的（按尾部数字取最大）。无则返回 None。"""
    if not rev:
        return None
    cands = glob.glob(os.path.join(DATA, base + "*.xlsx"))
    pref = [c for c in cands if rev in os.path.basename(c)]
    if not pref:
        return None

    def key(p):
        nums = re.findall(r"\d+", os.path.basename(p))
        return int(nums[-1]) if nums else 0

    return max(pref, key=key)


# ---------- 读取：偿付能力主表 ----------
def read_solvency(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr_i = None
    for ri, r in enumerate(rows):
        if r and any(str(c).strip() in SOLV_COL_MAP for c in r if c):
            hdr_i = ri
            break
    if hdr_i is None:
        print(f"  [WARN] 未找到表头: {os.path.basename(path)}")
        return {}
    headers = [str(c).strip() if c else "" for c in rows[hdr_i]]
    col_idx = {SOLV_COL_MAP[h]: ci for ci, h in enumerate(headers) if h in SOLV_COL_MAP}
    out = {}
    for ri in range(hdr_i + 1, len(rows)):
        r = rows[ri]
        if not r or not r[0]:
            continue
        comp = str(r[0]).strip()
        vals = {}
        for f, ci in col_idx.items():
            if ci < len(r) and r[ci] is not None:
                try:
                    v = float(r[ci])
                    if v == v:
                        vals[f] = round(v, 6)
                except Exception:
                    pass
        if vals:
            out[comp] = vals
    return out


# ---------- 读取：实际资本明细（横向，列A=指标，其余列=公司）----------
def read_capital(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    metric_row = {}
    for ri, r in enumerate(rows):
        name = r[0]
        if name is None:
            continue
        f = CAP_ROW_MAP.get(str(name).strip())
        if f:
            metric_row[f] = ri
    CAP_ORDER = ["core1", "net", "adj", "adj_nonrec", "adj_lti", "adj_invprop",
                 "adj_dta", "adj_cat", "adj_fvs", "adj_liab", "adj_other",
                 "core2", "c2_pref", "c2_fvm", "c2_oth", "c2_ded", "sub1", "sub2", "total"]
    out = {}
    for ci in range(1, len(rows[0])):
        comp = rows[0][ci]
        if comp is None:
            continue
        comp = str(comp).strip()
        rec = {}
        for f in CAP_ORDER:
            ri = metric_row.get(f)
            v = rows[ri][ci] if (ri is not None and ci < len(rows[ri])) else None
            if v is None:
                rec[f] = None
            else:
                try:
                    rec[f] = round(float(v), 6)
                except Exception:
                    rec[f] = None
        if any(v is not None for v in rec.values()):
            out[comp] = rec
    return out


# ---------- 读取：最低资本明细（横向，行0=公司/行1=期次/行N=指标）----------
def read_mc_excel(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header_row = rows[0]
    period_row = rows[1]

    def parse_period(s):
        m = re.match(r"(\d{4})第(\d)季度", str(s).strip())
        return f"{m.group(1)}Q{m.group(2)}" if m else None

    # 指标名 -> 行号
    agg_map = {}      # name -> field (P..N)
    detail_map = {}   # name -> field (mcP_loss..mcAdd_oth)
    fallback_agg = {} # row_idx -> field (旧版兜底)
    for ri, r in enumerate(rows):
        name = str(r[0]).strip() if r[0] is not None else ""
        if not name:
            continue
        if name in MC_AGG_MAP:
            agg_map[name] = MC_AGG_MAP[name]
        if name in MC_DETAIL_MAP:
            detail_map[name] = MC_DETAIL_MAP[name]
        if ri in MC_ROW_MAP:
            fallback_agg[ri] = MC_ROW_MAP[ri]

    use_name_agg = bool(agg_map)

    out = {}
    current_comp = None
    for col in range(1, len(header_row)):
        cell = header_row[col]
        if cell and str(cell).strip():
            current_comp = str(cell).strip()
        if not current_comp:
            continue
        pkey = parse_period(period_row[col])
        if not pkey:
            continue
        agg_vals = {}
        detail_vals = {}
        for row_idx, r in enumerate(rows):
            cv = r[col] if col < len(r) else None
            if cv in (None, ""):
                continue
            try:
                fv = round(float(cv), 6)
            except (ValueError, TypeError):
                continue
            name = str(r[0]).strip() if r[0] is not None else ""
            # 聚合项
            field = agg_map.get(name) if use_name_agg else fallback_agg.get(row_idx)
            if field:
                agg_vals[field] = fv
            # 明细项
            df = detail_map.get(name)
            if df:
                detail_vals[df] = fv
        if agg_vals or detail_vals:
            out.setdefault(current_comp, {})[pkey] = {"agg": agg_vals, "detail": detail_vals}
    return out


# ---------- 读取：监管披露行业平均 xls ----------
def read_reg_xls(path, quarter):
    import xlrd
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    qcol = None
    qname = {"Q1": "一季度末", "Q2": "二季度末", "Q3": "三季度末", "Q4": "四季度末"}[quarter[-2:]]
    hdr_row = None
    for r in range(ws.nrows):
        for c in range(ws.ncols):
            v = ws.cell(r, c).value
            if isinstance(v, str) and qname in v:
                hdr_row, qcol = r, c
                break
        if hdr_row is not None:
            break
    if qcol is None:
        print(f"  [WARN] 未在 {os.path.basename(path)} 找到 {qname} 列")
        return {}
    out = {}
    cur_metric = None
    for r in range(hdr_row + 1, ws.nrows):
        m0 = str(ws.cell(r, 0).value).strip()
        if m0:
            cur_metric = m0
        is_c = cur_metric and "综合偿付能力" in cur_metric
        is_d = cur_metric and "核心偿付能力" in cur_metric
        if not (is_c or is_d):
            continue
        seg_name = str(ws.cell(r, 1).value).strip()
        seg = REG_SEG_MAP.get(seg_name)
        if not seg:
            continue
        v = ws.cell(r, qcol).value
        if v in (None, ""):
            continue
        try:
            v = round(float(v) * 100.0, 2)
        except Exception:
            continue
        out.setdefault(seg, {})["C" if is_c else "D"] = round(v, 4)
    return out


# ---------- 通用 JS 读/写（保留 let + 头部注释）----------
def load_js(path):
    txt = open(path, encoding="utf-8").read()
    m = re.search(r"(let|const)\s+(\w+)\s*=\s*(\{.*\})\s*;", txt, re.S)
    if not m:
        raise SystemExit(f"[ERR] 解析失败: {path}")
    header = txt.split("\n", 1)[0]
    return m.group(2), json5.loads(m.group(3)), header


def save_js(path, var, obj, header, dry):
    js = header + "\nlet " + var + " = " + json.dumps(obj, ensure_ascii=False, separators=(",", ":")) + ";\n"
    if dry:
        print(f"  [DRY] 将写出 {os.path.basename(path)} ({len(js)//1024} KB)")
        return
    tmp = path + ".tmp"
    open(tmp, "w", encoding="utf-8").write(js)
    os.replace(tmp, path)  # 原子写，不留 .bak


# ============================================================
def upsert_datajs(quarter, rev, dry, overwrite, paths=None):
    print("=== data.js（偿付能力主表）===")
    var, D, header = load_js(DATA_JS)
    pdef = period_def(quarter)
    files = {"group": "集团偿付能力", "life": "人身险偿付能力",
             "property": "产险偿付能力", "reins": "再保偿付能力"}
    changed = False
    for seg, base in files.items():
        p = (paths or {}).get(seg) or rev_files(base, rev)
        if not p:
            print(f"  [SKIP] 无源文件: {base}")
            continue
        recs = read_solvency(p)
        segobj = D["segments"][seg]
        old = segobj["data"]
        n_new = n_ovw = n_skip = 0
        for comp, vals in recs.items():
            if comp not in segobj["companies"]:
                segobj["companies"].append(comp)
                changed = True
            od = old.setdefault(comp, {})
            if quarter in od and od[quarter] and any(od[quarter].values()):
                if overwrite:
                    od[quarter] = vals
                    n_ovw += 1
                    changed = True
                else:
                    n_skip += 1
            else:
                od[quarter] = vals
                n_new += 1
                changed = True
        if not any(pd["key"] == quarter for pd in segobj["periods"]):
            segobj["periods"].append(pdef)
            segobj["periods"].sort(key=lambda x: (x["year"], x["q"]))
            changed = True
        print(f"  {seg}: 新增 {n_new} / 覆盖 {n_ovw} / 跳过 {n_skip}  (源 {len(recs)} 家, {os.path.basename(p)})")
    if changed and not dry:
        save_js(DATA_JS, var, D, header, dry)
    elif dry:
        save_js(DATA_JS, var, D, header, dry)
    else:
        print("  data.js 无变更，跳过写文件")


def upsert_reg(quarter, rev, dry, quiet=False):
    if not quiet:
        print("=== reg_industry.js（监管行业平均）===")
    import glob as _glob
    yr = quarter[:4]
    cands = _glob.glob(os.path.join(DATA, "*监管披露行业偿付能力水平*.xls"))
    cands.sort(key=lambda x: (0 if yr in os.path.basename(x) else 1, x))
    p = cands[0] if cands else None
    if not p:
        if not quiet:
            print("  [SKIP] 未找到监管披露行业偿付能力水平*.xls")
        return
    recs = read_reg_xls(p, quarter)
    if not recs:
        if not quiet:
            print(f"  [SKIP] {os.path.basename(p)} 中无 {quarter} 数据")
        return
    var, R, header = load_js(REG_JS)
    n = 0
    for seg, cd in recs.items():
        if quarter in R["data"].get(seg, {}) and R["data"][seg][quarter]:
            continue
        R["data"].setdefault(seg, {})[quarter] = cd
        n += 1
        if not quiet:
            print(f"  {seg}: C={cd.get('C')} D={cd.get('D')}")
    if n > 0 and not dry:
        save_js(REG_JS, var, R, header, dry)
    elif dry and n > 0:
        save_js(REG_JS, var, R, header, dry)
    elif not quiet:
        print("  reg_industry.js 无变更，跳过写文件")


def upsert_capital(quarter, rev, dry, overwrite, paths=None):
    print("=== 实际资本明细 ===")
    for seg, base, js_path in [("life", "人身险实际资本", LIFE_CAP_JS),
                               ("property", "产险实际资本", PROP_CAP_JS)]:
        p = (paths or {}).get(seg) or rev_files(base, rev)
        if not p:
            print(f"  [SKIP] 无源文件: {base}")
            continue
        recs = read_capital(p)
        var, CAP, header = load_js(js_path)
        if quarter not in CAP["periods"]:
            CAP["periods"].append(quarter)
            CAP["periods"].sort()
            changed = True
        else:
            changed = False
        n_new = n_ovw = n_skip = 0
        for comp, rec in recs.items():
            if comp not in CAP["companies"]:
                CAP["companies"].append(comp)
                changed = True
            existing = CAP["data"].get(comp, {}).get(quarter)
            if existing and any(v is not None for v in existing.values()):
                if overwrite:
                    CAP["data"].setdefault(comp, {})[quarter] = rec
                    n_ovw += 1
                    changed = True
                else:
                    n_skip += 1
            else:
                CAP["data"].setdefault(comp, {})[quarter] = rec
                n_new += 1
                changed = True
        if changed and not dry:
            save_js(js_path, var, CAP, header, dry)
        elif dry:
            save_js(js_path, var, CAP, header, dry)
        else:
            print(f"  {seg}: 无变更，跳过写文件")
        print(f"  {seg}: 新增 {n_new} / 覆盖 {n_ovw} / 跳过 {n_skip}  (源 {len(recs)} 家, {os.path.basename(p)})")


def upsert_mc(quarter, rev, dry, overwrite, paths=None):
    print("=== 最低资本明细 (P~V+N + mcDetail 下钻) ===")
    mc_bases = {"life": "人身险最低资本", "property": "产险最低资本", "reins": "再保最低资本"}
    var, D, header = load_js(DATA_JS)
    changed = False
    for seg, base in mc_bases.items():
        p = (paths or {}).get(seg) or rev_files(base, rev)
        if not p:
            print(f"  [SKIP] 无源文件: {base}")
            continue
        recs = read_mc_excel(p)  # {comp: {period: {"agg":{P..N},"detail":{mcP_loss..}}}}}
        segobj = D["segments"][seg]
        mcDetail = segobj.setdefault("mcDetail", {})
        n_new = n_ovw = n_skip = 0
        d_new = d_ovw = d_skip = 0
        for comp, permap in recs.items():
            if quarter not in permap:
                continue
            vals = permap[quarter]
            if comp not in segobj["companies"]:
                print(f"  [WARN] {comp} 不在 data.js {seg} 公司列表，跳过")
                continue
            # 1) 主表聚合 P~V+N
            od = segobj["data"].setdefault(comp, {})
            rec = od.setdefault(quarter, {})
            has = any(k in rec for k in ("P", "Q", "R", "S", "T", "U", "V", "N"))
            if has and not overwrite:
                n_skip += 1
            else:
                for f, v in vals.get("agg", {}).items():
                    rec[f] = v
                if has:
                    n_ovw += 1
                else:
                    n_new += 1
                changed = True
            # 2) mcDetail 下钻明细
            det = vals.get("detail", {})
            if det:
                existing_det = mcDetail.setdefault(comp, {}).get(quarter)
                has_det = existing_det and any(v is not None for v in existing_det.values())
                if has_det and not overwrite:
                    d_skip += 1
                else:
                    mcDetail[comp][quarter] = det
                    if has_det:
                        d_ovw += 1
                    else:
                        d_new += 1
                    changed = True
        print(f"  {seg}: 主表 {n_new + n_ovw} 家(新{n_new}/覆盖{n_ovw}/跳过{n_skip})  mcDetail {d_new + d_ovw} 家(新{d_new}/覆盖{d_ovw}/跳过{d_skip})  ({os.path.basename(p)})")
    if changed and not dry:
        save_js(DATA_JS, var, D, header, dry)
    elif dry:
        save_js(DATA_JS, var, D, header, dry)
    else:
        print("  最低资本明细无变更，跳过写文件")


# ============================================================
def deploy(quarter):
    print("=== 部署 ===")
    for f in ["app.js", "index.html", "echarts.min.js", "compare_config.js",
              "life_capital_detail.js", "property_capital_detail.js",
              "data.js", "reg_industry.js"]:
        shutil.copy2(f, os.path.join("dist", f))
    subprocess.run([sys.executable, "_build_standalone.py"], check=True)
    shutil.copy2("偿付能力分析平台_可转发.html", "dist/偿付能力分析平台_可转发.html")
    subprocess.run(["git", "add", "-A"], check=True)
    subprocess.run(["git", "commit", "-m", f"data: upsert {quarter} (update_quarter.py)"], check=True)
    subprocess.run(["git", "push"], check=True)
    import urllib.request
    for url in ["https://yyseles.github.io/solvency/data.js",
                "https://yyseles.github.io/solvency/reg_industry.js"]:
        try:
            html = urllib.request.urlopen(url, timeout=20).read().decode("utf-8", "ignore")
            print(f"  线上 {url.split('/')[-1]} 含 {quarter}:", quarter in html)
        except Exception as e:
            print("  [WARN] 在线校验失败:", e)


# ============================================================
# 自动发现：扫描 data/ 下源文件，按文件名识别 板块类型+期次+修订号
# ============================================================
# 文件名前缀 -> (类型, 参数)
# 类型: solvency(偿付能力主表,data.js) / capital(实际资本明细) / mc(最低资本明细) / reg(监管披露)
AUTO_RULES = [
    ("集团偿付能力", "solvency", "group"),
    ("人身险偿付能力", "solvency", "life"),
    ("产险偿付能力", "solvency", "property"),
    ("再保偿付能力", "solvency", "reins"),
    ("人身险实际资本", "capital", "life"),
    ("产险实际资本", "capital", "property"),
    ("人身险最低资本", "mc", "life"),
    ("产险最低资本", "mc", "property"),
    ("再保最低资本", "mc", "reins"),
]

PERIOD_RE = [
    re.compile(r"(\d{4})Q([1-4])", re.I),          # 2026Q2 / 2026q2
    re.compile(r"(\d{4})年第?([1-4])季度?"),        # 2026年第2季度
    re.compile(r"(\d{2})Q([1-4])"),                 # 26Q2
    re.compile(r"(\d{2})年第?([1-4])季度?"),        # 26年第2季度
]


def parse_period_from_name(name):
    """从文件名提取期次键（2026Q2 或 2026）。无则返回 None。"""
    for rx in PERIOD_RE:
        m = rx.search(name)
        if m:
            y = int(m.group(1))
            if y < 100:
                y += 2000
            return f"{y}Q{m.group(2)}"
    m = re.search(r"(\d{4})年", name)              # 仅年份（如 2026年监管披露…）
    if m:
        return m.group(1)
    return None


def rev_of(name):
    """文件名尾部数字作为修订号（如 -0831 / 0831）。无则 0。"""
    nums = re.findall(r"\d+", os.path.basename(name))
    return int(nums[-1]) if nums else 0


def discover_files(data_dir, default_quarter, rev_only=None):
    """
    扫描 data/ 返回待处理清单：
      [{type, seg, quarter, rev, path}, ...]
    自动模式：识别所有带期次的文件；同一 (type,seg,quarter) 取修订号最大的一份。
    rev_only 指定时（手动模式）：只保留文件名含 rev_only 的文件。
    """
    tasks = []
    for p in sorted(glob.glob(os.path.join(data_dir, "*.xlsx"))):
        base = os.path.basename(p)
        if base.startswith("~$"):
            continue
        matched = None
        for prefix, typ, seg in AUTO_RULES:
            if base.startswith(prefix):
                matched = (typ, seg)
                break
        if not matched:
            continue
        if rev_only and rev_only not in base:
            continue
        q = parse_period_from_name(base)
        rv = rev_of(p)
        if not q:
            # 无期次标记：带修订号(如 -0831/-0918)的新文件 → 用默认期次；
            # 完全无修订号(纯文件名如“集团偿付能力.xlsx”) → 历史旧文件，跳过
            if rv > 0:
                q = default_quarter
                if not rev_only:
                    print(f"  [NOTE] 文件名无期次标记，按默认期次 {default_quarter} 处理: {base}")
            else:
                print(f"  [SKIP] 文件名无期次且无修订号(历史旧文件): {base}")
                continue
        tasks.append({"type": matched[0], "seg": matched[1],
                      "quarter": q, "rev": rv, "path": p})
    # 同 (type,seg,quarter) 只保留修订号最大的一份
    best = {}
    for t in tasks:
        key = (t["type"], t["seg"], t["quarter"])
        if key not in best or t["rev"] > best[key]["rev"]:
            best[key] = t
    tasks = list(best.values())
    tasks.sort(key=lambda t: (t["type"], t["seg"], t["quarter"]))
    return tasks


def auto_run(quarter, dry, overwrite, do_deploy, rev_only=None):
    """自动（或手动 rev）模式的统一入口：发现文件 → 按类型+期次分组 → 分发 → 汇总打印。"""
    tasks = discover_files(DATA, quarter, rev_only)
    if not tasks:
        print("[WARN] 未发现任何可更新的源文件。请把新季度 Excel 放入 data/ 目录。")
        return
    print(f"发现 {len(tasks)} 个待处理源文件：")
    for t in tasks:
        print(f"  [{t['type']:8s}] {t['seg']:8s} {t['quarter']}  rev={t['rev']:<5d} {os.path.basename(t['path'])}")
    print()

    # 按 (type, quarter) 分组：同一组一次调用，每个源文件只读一遍
    groups = {}
    for t in tasks:
        groups.setdefault((t["type"], t["quarter"]), {})[t["seg"]] = t["path"]

    n_solv = n_cap = n_mc = 0
    for (typ, q), segmap in sorted(groups.items()):
        if typ == "solvency":
            upsert_datajs(q, None, dry, overwrite, paths=segmap)
            n_solv += 1
        elif typ == "capital":
            upsert_capital(q, None, dry, overwrite, paths=segmap)
            n_cap += 1
        elif typ == "mc":
            upsert_mc(q, None, dry, overwrite, paths=segmap)
            n_mc += 1
        print()
    # 监管披露：扫描全部年份 xls，对每份文件逐季度尝试入库（已有数据自动跳过）
    import glob as _g
    for p in sorted(_g.glob(os.path.join(DATA, "*监管披露行业偿付能力水平*.xls"))):
        m = re.search(r"(20\d{2})年", os.path.basename(p))
        if not m:
            print(f"  [SKIP] 监管文件无法识别年份: {os.path.basename(p)}")
            continue
        yr = m.group(1)
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            upsert_reg(f"{yr}{q}", None, dry, quiet=True)

    print(f"处理完成：主表 {n_solv} 组 / 实际资本 {n_cap} 组 / 最低资本 {n_mc} 组。")
    if do_deploy and not dry:
        deploy(quarter)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--quarter", default="2026Q2",
                    help="目标期次；自动模式下用作“文件名无期次标记”文件的默认期次")
    ap.add_argument("--rev", default=None,
                    help="手动模式：只处理文件名含该修订号的文件；缺省=自动扫描全部")
    ap.add_argument("--dry", action="store_true")
    ap.add_argument("--deploy", action="store_true")
    ap.add_argument("--no-overwrite", dest="overwrite", action="store_false")
    ap.set_defaults(overwrite=True)
    args = ap.parse_args()

    auto_run(args.quarter, args.dry, args.overwrite, args.deploy, rev_only=args.rev)


if __name__ == "__main__":
    main()
