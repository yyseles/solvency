# -*- coding: utf-8 -*-
"""
保险行业偿付能力数据管线
- 抽取四个板块 (集团/产险/寿险/再保) 的 xlsx
- 口径归一：年末优先用"年度"(审计后)，否则用"Q4"(审计前)；季度 Q1/Q2/Q3 直接用
- 生成 solvency/data.js（按板块组织，全量内联）
- 支持后续更新：
    update --mode append|merge|patch  -> upsert：仅更新文件中出现的「公司×期间」，其余保持不变（修正某条错误记录用此）
    update --mode replace             -> 该板块整体由新文件替换

用法：
  python data_pipeline.py build  [--src DIR] [--out ../data.js]
  python data_pipeline.py update --segment life --xlsx 寿险偿付能力.xlsx --mode append  [--out ../data.js]
  python data_pipeline.py update --segment life --xlsx 寿险偿付能力.xlsx --mode replace [--out ../data.js]
"""
import argparse, json, os, re, sys
import openpyxl

# ---------- 板块定义 ----------
SEGMENTS = {
    'group':    {'name': '保险集团控股公司', 'file': '集团偿付能力.xlsx'},
    'property': {'name': '财产保险公司',     'file': '产险偿付能力.xlsx'},
    'life':     {'name': '人身保险公司',     'file': '寿险偿付能力.xlsx'},
    'reins':    {'name': '再保险公司',       'file': '再保偿付能力.xlsx'},
}
SEG_ORDER = ['group', 'property', 'life', 'reins']

# 最低资本明细文件（仅产/寿/再保；集团成员公司资本明细未披露，暂缺）
MC_FILES = {
    'life':     '寿险最低资本.xlsx',
    'property': '产险最低资本.xlsx',
    'reins':    '再保最低资本.xlsx',
}

# 最低资本明细 33 项指标（C-ROSS 二期标准模板），字段 key -> 行号
# 行号 = all_rows 的 0 索引下标（all_rows[0]=公司名行, all_rows[1]=期间行,
#   all_rows[2]=量化风险合计(行3), all_rows[3]=寿险合计(行4), ... all_rows[34]=最低资本(行35)）
# 其中与主料重合的 9 个聚合项(O/P/Q/R/S/T/U/V/N)不再重复存储
MC_FIELDS = [
    (2,'mcO'),(3,'mcP'),(4,'mcP_loss'),(5,'mcP_surr'),(6,'mcP_exp'),(7,'mcP_div'),
    (8,'mcQ'),(9,'mcQ_prem'),(10,'mcQ_cata'),(11,'mcQ_div'),
    (12,'mcR'),(13,'mcR_rate'),(14,'mcR_eq'),(15,'mcR_re'),(16,'mcR_ofb'),
    (17,'mcR_ofe'),(18,'mcR_fx'),(19,'mcR_div'),
    (20,'mcS'),(21,'mcS_spr'),(22,'mcS_def'),(23,'mcS_div'),
    (24,'mcT'),(25,'mcU'),(26,'mcU_nocap'),(27,'mcU_cap'),
    (28,'mcV'),(29,'mcAdd'),(30,'mcAdd_cyc'),(31,'mcAdd_dsii'),(32,'mcAdd_gsii'),
    (33,'mcAdd_oth'),(34,'mcN'),
]
# 主料已含聚合项（万元同单位），明细仅存"进一步细分"子项 + 附加资本(主料无)
MC_EXCLUDE = {'mcO','mcP','mcQ','mcR','mcS','mcT','mcU','mcV','mcN'}

COLMAP = {2:'C',3:'D',4:'E',5:'F',6:'G',7:'H',8:'I',9:'J',10:'K',11:'L',12:'M',
          13:'N',14:'O',15:'P',16:'Q',17:'R',18:'S',19:'T',20:'U',21:'V'}
LABELS = {'C':'综合偿付能力充足率','D':'核心偿付能力充足率','E':'综合偿付能力溢额',
 'F':'核心偿付能力溢额','G':'认可资产','H':'认可负债','I':'实际资本','J':'核心一级资本',
 'K':'核心二级资本','L':'附属一级资本','M':'附属二级资本','N':'最低资本',
 'O':'量化风险最低资本','P':'寿险风险最低资本','Q':'非寿险风险最低资本',
 'R':'市场风险最低资本','S':'信用风险最低资本','T':'量化风险分散效应',
 'U':'特定类别保险合同损失吸收效应','V':'控制风险最低资本'}

def parse_period(s):
    """返回 (year, q, kind) 或 None。kind: 'Q1'|'Q2'|'Q3'|'Q4'|'YR'"""
    if not s: return None
    m = re.match(r'(\d{4})年?(第一季度|第二季度|第三季度|第四季度|年度)', s)
    if not m: return None
    y = int(m.group(1)); seg = m.group(2)
    return {'第一季度':(y,1,'Q1'),'第二季度':(y,2,'Q2'),'第三季度':(y,3,'Q3'),
            '第四季度':(y,4,'Q4'),'年度':(y,4,'YR')}[seg]

def extract_segment(xlsx_path):
    """从单个 xlsx 抽取并归一化，返回 {companies, periods, data}"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    # raw[comp][rawKey] = rec ;  rawKey: "2024YR" / "2024Q4" / "2024Q1"
    raw = {}
    companies = set()
    years = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        comp = row[0]; per = row[1]
        if comp is None or per is None: continue
        comp = str(comp).strip()
        if not comp: continue
        pp = parse_period(per)
        if not pp: continue
        y, q, kind = pp
        years.add(y)
        companies.add(comp)
        rawKey = f"{y}{kind}"
        rec = {}
        for idx, k in COLMAP.items():
            v = row[idx]
            if v is None: continue
            if isinstance(v, str):
                try: v = float(v)
                except: continue
            if isinstance(v, float) and (v != v): continue  # NaN
            rec[k] = v
        raw.setdefault(comp, {})[rawKey] = rec
    wb.close()

    # 归一化：每年年末优先 YR 否则 Q4；季度 Q1/Q2/Q3 直接用
    data = {}
    period_meta = {}  # normKey -> dict
    for comp, rmap in raw.items():
        out = {}
        for y in sorted(years):
            # 年末
            if f"{y}YR" in rmap:
                nk = f"{y}"; src = 'audited'
                out[nk] = rmap[f"{y}YR"]
                period_meta[nk] = {'key':nk,'year':y,'q':4,'kind':'year-end',
                                   'label':f"{y}年",'source':src}
            elif f"{y}Q4" in rmap:
                nk = f"{y}"; src = 'prelim'
                out[nk] = rmap[f"{y}Q4"]
                period_meta[nk] = {'key':nk,'year':y,'q':4,'kind':'year-end',
                                   'label':f"{y}年",'source':src}
            # 季度
            for q in (1,2,3):
                rk = f"{y}Q{q}"
                if rk in rmap:
                    out[rk] = rmap[rk]
                    period_meta[rk] = {'key':rk,'year':y,'q':q,'kind':'quarter',
                                       'label':f"{y}年第{q}季度",'source':'quarter'}
        if out:
            data[comp] = out

    periods = sorted(period_meta.values(),
                     key=lambda p:(p['year'], {1:1,2:2,3:3,4:4}[p['q']]))
    return {'companies': sorted(data.keys()), 'periods': periods, 'data': data}

def extract_mc(xlsx_path):
    """从最低资本明细宽表（行=指标，列=公司×20期）抽取。
    返回 {company: {periodKey: {field: value(万元)}}}，仅存 MC_EXCLUDE 之外的细分项。
    periodKey: 202X第1-3季度 -> 202XQ1..Q3；第4季度 -> 202X（与主料年末key一致）。
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    r1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    r2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    # 第1公司块(列1..20)的期间映射
    period_keys = []
    for j in range(20):
        s = r2[1+j] if (1+j) < len(r2) else None
        if s is None:
            period_keys.append(None); continue
        m = re.match(r'(\d{4})第([1-4])季度', str(s).strip())
        if not m:
            period_keys.append(None); continue
        y = int(m.group(1)); q = int(m.group(2))
        period_keys.append(f"{y}" if q == 4 else f"{y}Q{q}")
    # 一次性读入全表（35行×1861列），避免 read_only 模式下单行随机访问的不稳定
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # 行索引：0=指标名称行, 1=期间行, 2..34=33个指标行
    # 公司块起始列：1,21,41,...
    ncols = len(r1)
    result = {}
    for ci in range(1, ncols, 20):
        comp = r1[ci]
        if comp is None: continue
        comp = str(comp).strip()
        out = {}
        for (ri, field) in MC_FIELDS:
            if field in MC_EXCLUDE: continue
            rowvals = all_rows[ri] if ri < len(all_rows) else None
            if rowvals is None: continue
            for j in range(20):
                pk = period_keys[j]
                if pk is None: continue
                col = ci + j
                v = rowvals[col] if col < len(rowvals) else None
                if v is None: continue
                if isinstance(v, str):
                    try: v = float(v)
                    except: continue
                if isinstance(v, float) and v != v: continue
                out.setdefault(pk, {})[field] = v
        if out: result[comp] = out
    return result

def attach_mc(seg, seg_obj, src_dir):
    """把最低资本明细挂接到 seg_obj（仅保留主料已有的期间）。无明细文件则置空。"""
    f = os.path.join(src_dir, MC_FILES.get(seg, ''))
    if seg not in MC_FILES or not os.path.exists(f):
        seg_obj['mcDetail'] = {}
        if seg in MC_FILES:
            print(f"  [mc] {seg}: 未找到 {MC_FILES[seg]}，mcDetail 置空")
        return
    mc = extract_mc(f)
    main_keys = {p['key'] for p in seg_obj['periods']}
    filtered = {}
    for comp, d in mc.items():
        fd = {k: v for k, v in d.items() if k in main_keys}
        if fd: filtered[comp] = fd
    seg_obj['mcDetail'] = filtered
    nrec = sum(len(v) for v in filtered.values())
    print(f"  [mc] {seg}: 挂接明细 {len(filtered)}公司/{nrec}条记录（已按主料期间过滤）")

def build_all(src_dir, out_path):
    seg_objs = {}
    for seg in SEG_ORDER:
        f = os.path.join(src_dir, SEGMENTS[seg]['file'])
        if not os.path.exists(f):
            print(f"[WARN] 缺少 {f}", file=sys.stderr); continue
        print(f"[build] {seg} <- {SEGMENTS[seg]['file']}")
        seg_objs[seg] = extract_segment(f)
    for seg in SEG_ORDER:
        if seg in seg_objs:
            attach_mc(seg, seg_objs[seg], src_dir)
    write_data(seg_objs, out_path)
    # 诊断
    for seg in SEG_ORDER:
        if seg not in seg_objs: continue
        o = seg_objs[seg]
        last = o['periods'][-1]['key'] if o['periods'] else '?'
        ye = [p for p in o['periods'] if p['kind']=='year-end']
        prelim = [p for p in ye if p['source']=='prelim']
        mc = o.get('mcDetail', {})
        print(f"  {seg}: {len(o['companies'])}公司, {len(o['periods'])}期, "
              f"最新={last}, 年末点{len(ye)}个(审计前{prelim}个), 明细{len(mc)}公司")

def load_existing(out_path):
    """读取现有 data.js 返回 dict（兼容旧格式）"""
    if not os.path.exists(out_path): return None
    txt = open(out_path, 'r', encoding='utf-8').read()
    m = re.search(r'=\s*(\{.*\})\s*;', txt, re.S)
    if not m: return None
    try: return json.loads(m.group(1))
    except Exception as e:
        print(f"[ERR] 解析现有 data.js 失败: {e}", file=sys.stderr); return None

def write_data(seg_objs, out_path):
    meta = {
        'unit': '万元',
        'ratioCols': ['C','D'],
        'labels': LABELS,
        'segments': SEG_ORDER,
        'segmentNames': {s: SEGMENTS[s]['name'] for s in SEG_ORDER},
        'caliber': '年末口径优先取"年度"(审计后)，无年度则取"Q4"(审计前)；季度取 Q1/Q2/Q3。不再区分双口径。',
    }
    out = {'meta': meta, 'segments': seg_objs}
    js = "// 保险行业偿付能力分析平台 - 数据文件（自动生成，勿手改；更新请运行 data_pipeline.py）\n"
    js += "const SOLVENCY_DATA = " + json.dumps(out, ensure_ascii=False) + ";\n"
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(js)
    print(f"[ok] 写出 {out_path}  ({len(js)//1024} KB)")

def update(seg, xlsx_path, mode, out_path):
    if seg not in SEGMENTS:
        print(f"[ERR] 未知板块 {seg}", file=sys.stderr); sys.exit(1)
    if not os.path.exists(xlsx_path):
        print(f"[ERR] 缺少 {xlsx_path}", file=sys.stderr); sys.exit(1)
    print(f"[update] segment={seg} mode={mode} xlsx={xlsx_path}")
    new_seg = extract_segment(xlsx_path)
    cur = load_existing(out_path)
    if cur is None:
        # 没有现有数据，等价于全量 build 该板块
        print("  (无现有 data.js，按全量写入)")
        write_data({seg: new_seg}, out_path)
        return
    segs = cur.get('segments', {})
    if mode == 'replace':
        segs[seg] = new_seg
        print(f"  replace: 板块 {seg} 已整体替换为新数据（{len(new_seg['companies'])}公司 {len(new_seg['periods'])}期）")
    else:  # append / merge / patch：upsert——仅更新文件中出现的「公司×期间」，其余保持不变
        old = segs.get(seg, {'companies':[], 'periods':[], 'data':{}})
        old_data = old.get('data', {})
        old_companies = set(old.get('companies', []))
        old_periods = {p['key']: p for p in old.get('periods', [])}
        n_upd = 0; n_new = 0
        for comp, pmap in new_seg['data'].items():
            old_companies.add(comp)
            od = old_data.setdefault(comp, {})
            for pk, rec in pmap.items():
                if pk in od: n_upd += 1
                else: n_new += 1
                od[pk] = rec
                # 更新期间 meta（以新文件为准）
                pm = next((p for p in new_seg['periods'] if p['key']==pk), None)
                if pm: old_periods[pk] = pm
        periods = sorted(old_periods.values(),
                         key=lambda p:(p['year'], {1:1,2:2,3:3,4:4}[p['q']]))
        segs[seg] = {'companies': sorted(old_companies), 'periods': periods, 'data': old_data}
        print(f"  upsert: 更新 {n_upd} 条、新增 {n_new} 条；板块 {seg} 现 {len(old_companies)}公司 {len(periods)}期（未出现在文件中的记录保持不变）")
    # 同步重挂最低资本明细（按更新后的期间过滤，保持与主数据一致）
    attach_mc(seg, segs[seg], os.path.dirname(os.path.abspath(xlsx_path)))
    write_data(segs, out_path)

def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest='cmd')
    pb = sub.add_parser('build')
    pb.add_argument('--src', default=os.path.dirname(os.path.abspath(__file__))+'/../data')
    pb.add_argument('--out', default=os.path.dirname(os.path.abspath(__file__))+'/../data.js')
    pu = sub.add_parser('update')
    pu.add_argument('--segment', required=True, choices=SEG_ORDER)
    pu.add_argument('--xlsx', required=True)
    pu.add_argument('--mode', required=True, choices=['append','merge','patch','replace'])
    pu.add_argument('--out', default=os.path.dirname(os.path.abspath(__file__))+'/../data.js')
    args = ap.parse_args()
    if args.cmd == 'build':
        build_all(args.src, args.out)
    elif args.cmd == 'update':
        update(args.segment, args.xlsx, args.mode, args.out)
    else:
        ap.print_help()

if __name__ == '__main__':
    main()
