# 临时脚本：按 0831 模板生成「中再集团本级」26Q2 的再保偿付能力/最低资本 xlsx（仅1家公司）
import openpyxl

COMP = "中再集团本级"

# ---- 偿付能力主表数值（万元；C/D 为小数比率） ----
solv_vals = {
    "综合偿付能力充足率": 2.3534,
    "核心偿付能力充足率": 2.3471,
    "综合偿付能力溢额": 6408927.55,
    "核心偿付能力溢额": 6378885.99,
    "认可资产": 15578358.79,
    "认可负债": 4434094.37,
    "实际资本": 11144264.42,
    "核心一级资本": 11114222.86,
    "核心二级资本": 0.0,
    "附属一级资本": 30041.56,
    "附属二级资本": 0.0,
    "最低资本": 4735336.87,
    "量化风险最低资本": 4726803.11,
    "寿险风险最低资本": 16736.89,
    "非寿险风险最低资本": 1251927.05,
    "市场风险最低资本": 4674088.44,
    "信用风险最低资本": 63881.58,
    "量化风险分散效应": 1024085.62,
    "特定类别保险合同损失吸收效应": 6966.12,
    "控制风险最低资本": 8533.76,
}

# ---- 最低资本明细数值（万元） ----
mc_vals = {
    "量化风险最低资本": 4726803.11,
    "寿险业务保险风险最低资本合计": 16736.89,
    "寿险业务保险风险-损失发生风险最低资本": 15329.34,
    "寿险业务保险风险-退保风险最低资本": 5875.95,
    "寿险业务保险风险-费用风险最低资本": 567.08,
    "寿险业务保险风险-风险分散效应": 5035.48,
    "非寿险业务保险风险最低资本合计": 1251927.05,
    "非寿险业务保险风险-保费及准备金风险最低资本": 780736.03,
    "非寿险业务保险风险-巨灾风险最低资本": 802748.56,
    "非寿险业务保险风险-风险分散效应": 331557.53,
    "市场风险-最低资本合计": 4674088.44,
    "市场风险-利率风险最低资本": 208439.80,
    "市场风险-权益价格风险最低资本": 4216158.47,
    "市场风险-房地产价格风险最低资本": 11691.60,
    "市场风险-境外固定收益类资产价格风险最低资本": 75890.66,
    "市场风险-境外权益类资产价格风险最低资本": 839249.81,
    "市场风险-汇率风险最低资本": 93375.11,
    "市场风险-风险分散效应": 770717.02,
    "信用风险-最低资本合计": 63881.58,
    "信用风险-利差风险最低资本": 12399.71,
    "信用风险-交易对手违约风险最低资本": 59643.30,
    "信用风险-风险分散效应": 8161.43,
    "量化风险分散效应": 1024085.62,
    "特定类别保险合同损失吸收效应": 6966.12,
    "损失吸收调整-不考虑上限": 15426.33,
    "损失吸收效应调整上限": 6966.12,
    "控制风险最低资本": 8533.76,
    "附加资本": 0.0,
    "最低资本": 4735336.87,
}

# ===== 偿付能力主表 =====
src = openpyxl.load_workbook("data/再保偿付能力26Q2-0831.xlsx", data_only=True)
ws = src.active
wb = openpyxl.Workbook()
nw = wb.active
nw.title = ws.title
for c in range(1, ws.max_column + 1):
    nw.cell(row=1, column=c, value=ws.cell(row=1, column=c).value)
    nw.cell(row=2, column=c, value=ws.cell(row=2, column=c).value)
nw.cell(row=3, column=1, value=COMP)
nw.cell(row=3, column=2, value="2026第二季度")
for c in range(3, ws.max_column + 1):
    lab = ws.cell(row=2, column=c).value
    if lab in solv_vals:
        nw.cell(row=3, column=c, value=solv_vals[lab])
wb.save("data/再保偿付能力26Q2-0901.xlsx")
print("写出生效: 再保偿付能力26Q2-0901.xlsx")

# ===== 最低资本明细 =====
msrc = openpyxl.load_workbook("data/再保最低资本26Q2-0831.xlsx", data_only=True)
mws = msrc["最低资本"]
mwb = openpyxl.Workbook()
mnw = mwb.active
mnw.title = "最低资本"
mnw.cell(row=1, column=1, value="指标名称")
mnw.cell(row=1, column=2, value=COMP)
mnw.cell(row=2, column=2, value="2026第2季度")
for r in range(3, mws.max_row + 1):
    lab = mws.cell(row=r, column=1).value
    mnw.cell(row=r, column=1, value=lab)
    if lab in mc_vals and mc_vals[lab] is not None:
        mnw.cell(row=r, column=2, value=mc_vals[lab])
mwb.save("data/再保最低资本26Q2-0901.xlsx")
print("写出生效: 再保最低资本26Q2-0901.xlsx")
