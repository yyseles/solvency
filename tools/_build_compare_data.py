import openpyxl, json, os

TOOLS_DIR = os.path.dirname(os.path.abspath(__file__))
SOLV_DIR = os.path.join(TOOLS_DIR, "..")
DATA_DIR = os.path.join(SOLV_DIR, "data")

EXCEL = os.path.join(DATA_DIR, '上市公司及其他主要公司对比.xlsx')
DATAJS = os.path.join(SOLV_DIR, 'data.js')

periods = ['22Q1','22Q2','22Q3','22Q4','23Q1','23Q2','23Q3','23Q4','24Q1','24Q2','24Q3','24Q4','25Q1','25Q2','25Q3','25Q4','26Q1']
col0 = 3  # column C

wb = openpyxl.load_workbook(EXCEL, read_only=False, data_only=True)
ws = wb.active

METRICS = ['综合偿付能力充足率','核心偿付能力充足率','实际资本','核心资本','附属资本','最低资本']
RATIO_METRICS = {'综合偿付能力充足率':'实际资本','核心偿付能力充足率':'核心资本'}

raw = {}  # raw[metric][block][company] = {period:val}
current_metric = None
current_block = None

for r in range(1, ws.max_row+1):
    a = ws.cell(row=r, column=1).value
    b = ws.cell(row=r, column=2).value
    if a in METRICS:
        current_metric = a
        if b in ['集团','寿险','产险','银行系主要公司']:
            current_block = b
        continue
    if a in ['集团','寿险','产险','银行系主要公司'] and b:
        current_block = a
        vals = {}
        for i, p in enumerate(periods):
            v = ws.cell(row=r, column=col0+i).value
            if v is not None and v != '':
                vals[p] = float(v)
        raw.setdefault(current_metric, {}).setdefault(current_block, {})[b] = vals
        continue
    if b and a is None:
        vals = {}
        for i, p in enumerate(periods):
            v = ws.cell(row=r, column=col0+i).value
            if v is not None and v != '':
                vals[p] = float(v)
        raw.setdefault(current_metric, {}).setdefault(current_block, {})[b] = vals

# --- 上市平均/合计 (剔除阳光系) ---
def excl_sun(companies):
    return [c for c in companies if '阳光' not in c]

agg = {}
for metric in METRICS:
    agg[metric] = {}
    for block in ['集团','寿险','产险']:
        comps = list(raw.get(metric, {}).get(block, {}).keys())
        incl = excl_sun(comps)
        if metric in RATIO_METRICS:
            num = RATIO_METRICS[metric]
            sets = {'上市公司平均-含众安': incl, '上市公司平均-不含众安': [c for c in incl if '众安' not in c]} if block=='产险' else {'上市公司平均': incl}
            agg[metric][block] = {}
            for label, cl in sets.items():
                series = {}
                for p in periods:
                    nsum = sum(raw.get(num, {}).get(block, {}).get(c, {}).get(p, 0) or 0 for c in cl)
                    dsum = sum(raw.get('最低资本', {}).get(block, {}).get(c, {}).get(p, 0) or 0 for c in cl)
                    series[p] = round(nsum/dsum, 6) if dsum else None
                agg[metric][block][label] = series
        else:
            sets = {'上市公司合计-含众安': incl, '上市公司合计-不含众安': [c for c in incl if '众安' not in c]} if block=='产险' else {'上市公司合计': incl}
            agg[metric][block] = {}
            for label, cl in sets.items():
                series = {}
                for p in periods:
                    ssum = sum(raw.get(metric, {}).get(block, {}).get(c, {}).get(p, 0) or 0 for c in cl)
                    series[p] = round(ssum, 4) if ssum else (0.0 if any(raw.get(metric,{}).get(block,{}).get(c,{}).get(p) is not None for c in cl) else None)
                agg[metric][block][label] = series

# --- 阳光系合计 (阳光集团/阳光人寿/阳光财产 跨板块加权) ---
sun_map = {'集团':'阳光集团','寿险':'阳光人寿','产险':'阳光财产'}
sunAgg = {}
for metric in METRICS:
    series = {}
    for p in periods:
        if metric in RATIO_METRICS:
            num = RATIO_METRICS[metric]
            nsum = sum(raw.get(num, {}).get(blk, {}).get(co, {}).get(p, 0) or 0 for blk, co in sun_map.items())
            dsum = sum(raw.get('最低资本', {}).get(blk, {}).get(co, {}).get(p, 0) or 0 for blk, co in sun_map.items())
            series[p] = round(nsum/dsum, 6) if dsum else None
        else:
            ssum = sum(raw.get(metric, {}).get(blk, {}).get(co, {}).get(p, 0) or 0 for blk, co in sun_map.items())
            series[p] = round(ssum, 4) if ssum else (0.0 if any(raw.get(metric,{}).get(blk,{}).get(co,{}).get(p) is not None for blk,co in sun_map.items()) else None)
    sunAgg[metric] = series

# --- 银行系 (仅人身险板块, 9家) 全指标 ---
banks = ['工银安盛','光大永明','建信人寿','交银人寿','中荷人寿','农银人寿','中银三星','中邮人寿','招商信诺']
with open(DATAJS, 'r', encoding='utf-8') as f:
    t = f.read()
i = t.index('SOLVENCY_DATA = ') + len('SOLVENCY_DATA = ')
depth = 0; e = 0
for k, c in enumerate(t[i:]):
    if c == '{': depth += 1
    elif c == '}':
        depth -= 1
        if depth == 0:
            e = k+1; break
SOL = json.loads(t[i:i+e])
life = SOL['segments']['life']

def dnkey(p):
    return ('20'+p[:2]) if p.endswith('Q4') else ('20'+p)

# data.js 字段: 分子字段(可多个求和), 分母字段(仅比率用)
FIELD = {
    '综合偿付能力充足率': (['I'], ['N']),
    '核心偿付能力充足率': (['J','K'], ['N']),
    '实际资本': (['I'], None),
    '核心资本': (['J','K'], None),
    '附属资本': (['L','M'], None),
    '最低资本': (['N'], None),
}

def gv(bank, dk, fld):
    return life['data'].get(bank, {}).get(dk, {}).get(fld) or 0

bankAgg = {}
bankRaw = {}
for metric in METRICS:
    numf, denf = FIELD[metric]
    is_ratio = denf is not None
    agg_series = {}
    per_company = {b: {} for b in banks}
    for p in periods:
        dk = dnkey(p)
        nsum = sum(sum(gv(b, dk, f) for f in numf) for b in banks)
        if is_ratio:
            dsum = sum(sum(gv(b, dk, f) for f in denf) for b in banks)
            agg_series[p] = round(nsum/dsum, 6) if dsum else None
            for b in banks:
                cn = sum(gv(b, dk, f) for f in numf)
                cd = sum(gv(b, dk, f) for f in denf)
                per_company[b][p] = round(cn/cd, 6) if cd else None
        else:
            # data.js 金额字段单位为万元，需÷10000转为亿元
            agg_series[p] = round(nsum / 10000, 2) if nsum else 0.0
            for b in banks:
                per_company[b][p] = round(sum(gv(b, dk, f) for f in numf) / 10000, 2)
    bankAgg[metric] = agg_series
    bankRaw[metric] = per_company

# --- 集团加权平均(计算口径): 所有集团公司加权(含阳光集团) ---
groupCalc = {}
for metric in METRICS:
    if metric in RATIO_METRICS:
        num = RATIO_METRICS[metric]
        series = {}
        for p in periods:
            nsum = sum((raw.get(num, {}).get('集团', {})[c] or {}).get(p, 0) or 0 for c in raw.get(num, {}).get('集团', {}))
            dsum = sum((raw.get('最低资本', {}).get('集团', {})[c] or {}).get(p, 0) or 0 for c in raw.get('最低资本', {}).get('集团', {}))
            series[p] = round(nsum/dsum, 6) if dsum else None
        groupCalc[metric] = series
    else:
        series = {}
        for p in periods:
            ssum = sum((raw.get(metric, {}).get('集团', {})[c] or {}).get(p, 0) or 0 for c in raw.get(metric, {}).get('集团', {}))
            series[p] = round(ssum, 4) if ssum else None  # 集团按半年报披露，Q1/Q3 无数据留空
        groupCalc[metric] = series

# --- 各板块明细公司列表(取自 Excel raw, 跨指标一致) ---
comp_list = {b: list(raw.get('综合偿付能力充足率', {}).get(b, {}).keys()) for b in ['集团','寿险','产险']}

# --- 按板块组织的对比实体 ---
sections = {
    'group': {
        'title': '保险集团', 'block': '集团',
        'entities': [
            {'name': '集团加权平均(计算口径)', 'src': 'calc_group'},
            {'name': '上市集团加权平均', 'src': 'agg', 'block': '集团', 'label': '上市公司平均'},
            {'name': '阳光集团', 'src': 'sun', 'block': '集团', 'company': '阳光集团'},
        ],
        'companies': comp_list['集团'],
    },
    'life': {
        'title': '人身险', 'block': '寿险',
        'entities': [
            {'name': '人身险行业加权平均(监管披露口径)', 'src': 'reg', 'seg': 'life'},
            {'name': '上市人身险公司平均', 'src': 'agg', 'block': '寿险', 'label': '上市公司平均'},
            {'name': '银保系平均', 'src': 'bank'},
            {'name': '阳光人寿', 'src': 'sun', 'block': '寿险', 'company': '阳光人寿'},
        ],
        'companies': comp_list['寿险'],
    },
    'property': {
        'title': '财产险', 'block': '产险',
        'entities': [
            {'name': '财险行业加权平均(监管披露口径)', 'src': 'reg', 'seg': 'property'},
            {'name': '上市财险平均(含众安)', 'src': 'agg', 'block': '产险', 'label': '上市公司平均-含众安'},
            {'name': '上市财险平均(不含众安)', 'src': 'agg', 'block': '产险', 'label': '上市公司平均-不含众安'},
            {'name': '阳光财险', 'src': 'sun', 'block': '产险', 'company': '阳光财产'},
        ],
        'companies': comp_list['产险'],
    },
}

out = {
    'periods': periods,
    'metrics': METRICS,
    'ratioMetrics': RATIO_METRICS,
    'sunMap': sun_map,
    'bankCompanies': banks,
    'raw': raw,
    'agg': agg,
    'sunAgg': sunAgg,
    'bankAgg': bankAgg,
    'bankRaw': bankRaw,
    'groupCalc': groupCalc,
    'sections': sections,
}

with open(os.path.join(SOLV_DIR, 'compare_data.js'), 'w', encoding='utf-8') as f:
    f.write('window.COMPARE_DATA = ')
    json.dump(out, f, ensure_ascii=False, indent=1)
    f.write(';\n')

# ---- validation ----
print('=== 校验: 重新计算的 上市公司平均 vs Excel 已知值 ===')
checks = [
    ('综合偿付能力充足率','集团','上市公司平均','22Q2', 2.4658679514),
    ('综合偿付能力充足率','寿险','上市公司平均','22Q2', 2.4110771124),
    ('综合偿付能力充足率','产险','上市公司平均-含众安','22Q2', 2.3247704024),
    ('综合偿付能力充足率','产险','上市公司平均-不含众安','22Q2', 2.3036258282),
    ('核心偿付能力充足率','集团','上市公司平均','22Q2', 1.8007589538),
    ('核心偿付能力充足率','产险','上市公司平均-不含众安','22Q2', 1.9303909636),
]
ok = True
for m, b, lbl, p, exp in checks:
    got = agg[m][b][lbl][p]
    d = abs(got-exp)
    flag = 'OK' if d < 1e-4 else 'FAIL'
    if d >= 1e-4: ok = False
    print('  %s/%s/%s/%s: got=%.10f exp=%.10f %s' % (m,b,lbl,p,got,exp,flag))
print('ALL OK' if ok else 'SOME FAILED')

print('\n=== 阳光系合计(加权) 抽样 ===')
for p in ['22Q1','22Q4','25Q4','26Q1']:
    print('  综合 %s: %.4f | 核心 %s: %.4f' % (p, sunAgg['综合偿付能力充足率'][p], p, sunAgg['核心偿付能力充足率'][p]))

print('\n=== 银行系合计(加权) 抽样 ===')
for p in ['22Q1','22Q4','25Q4','26Q1']:
    print('  综合 %s: %.4f | 核心 %s: %.4f' % (p, bankAgg['综合偿付能力充足率'][p], p, bankAgg['核心偿付能力充足率'][p]))

# 银行系核心单家抽样: 工银安盛 22Q1 应≈1.1527
print('\n  工银安盛 核心 22Q1:', bankRaw['核心偿付能力充足率']['工银安盛']['22Q1'], '(期望~1.1527)')

print('\n=== 银行系金额(亿元) 抽样 ===')
for p in ['22Q1','26Q1']:
    print('  实际资本合计 %s: %.2f' % (p, bankAgg['实际资本'][p]))
    print('  工银安盛 实际资本 %s: %.2f' % (p, bankRaw['实际资本']['工银安盛'][p]))

print('\n=== 集团加权平均(计算口径, 含阳光) 抽样 ===')
gc = groupCalc['综合偿付能力充足率']
none_ps = [p for p in periods if gc.get(p) is None]
print('  groupCalc 综合 为 None 的期次:', none_ps if none_ps else '无')
for p in ['22Q1','22Q2','25Q4','26Q1']:
    gv = gc.get(p); av = agg['综合偿付能力充足率']['集团']['上市公司平均'].get(p)
    print('  综合 %s: 计算口径=%s | 上市平均=%.4f' % (p, ('%.4f'%gv) if gv is not None else 'None', av if av is not None else 0))

print('\n=== 板块实体结构 ===')
for k, s in sections.items():
    print('  [%s] %s | 明细公司数=%d | 实体=%s' % (k, s['title'], len(s['companies']), [e['name'] for e in s['entities']]))

print('\ncompare_data.js bytes=%d' % len(json.dumps(out, ensure_ascii=False)))
