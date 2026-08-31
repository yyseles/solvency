import openpyxl, os, re

# 板块 -> (偿付能力表, 最低资本表)
PAIRS = [
    ('集团', '集团偿付能力.xlsx', None),
    ('人身险', '人身险偿付能力.xlsx', '人身险最低资本.xlsx'),
    ('产险', '产险偿付能力.xlsx', '产险最低资本.xlsx'),
    ('再保', '再保偿付能力.xlsx', '再保最低资本.xlsx'),
]

CN = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9,'十':10}
def norm_period(s):
    if not s: return None
    s = str(s).strip()
    m = re.search(r'(\d{4})', s)
    if not m: return None
    y = m.group(1)
    if '年度' in s or (len(s) == 4 and s.isdigit()):
        return y
    # 季度：支持 第一季度/第1季度/Q1
    qm = re.search(r'第([一二三四五六七八九十\d])季度|Q(\d)', s)
    if qm:
        qraw = qm.group(1) or qm.group(2)
        q = CN.get(qraw, qraw) if qraw else None
        if q: return f'{y}Q{q}'
    return y

# 主表：行=公司+期次, col0=公司, col1=期次, col13=最低资本
def read_main(f):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    res = {}  # (公司, 期次) -> 最低资本
    for r in rows[2:]:
        if not r or not r[0]: continue
        comp = str(r[0]).strip()
        per = norm_period(r[1]) if len(r) > 1 else None
        if per is None: continue
        val = r[13] if len(r) > 13 else None
        res[(comp, per)] = val
    return res

# 明细表：row0=表头(公司名), row1=期次, 指标行找"最低资本"汇总行
def read_detail(f):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = rows[0]  # 公司名表头
    per_row = rows[1]  # 期次
    # 找"最低资本"汇总行(不是"最低资本合计"子类)
    detail_row = None
    for ri, r in enumerate(rows[2:], start=2):
        name = str(r[0]).strip() if r and r[0] else ''
        if name == '最低资本':
            detail_row = ri
            break
    if detail_row is None:
        # 退而求其次 含"最低资本"且非子类的
        for ri, r in enumerate(rows[2:], start=2):
            name = str(r[0]).strip() if r and r[0] else ''
            if '最低资本' in name and '风险' not in name and '分散' not in name and '调整' not in name and '吸收' not in name and '附加' not in name and '成员' not in name:
                detail_row = ri
                break
    res = {}
    if detail_row is None:
        return res
    for ci in range(1, len(hdr)):
        comp = str(hdr[ci]).strip() if hdr[ci] else ''
        if not comp or comp == 'None': continue
        per = norm_period(per_row[ci]) if ci < len(per_row) else None
        if per is None: continue
        val = rows[detail_row][ci] if ci < len(rows[detail_row]) else None
        res[(comp, per)] = val
    return res

print('=' * 80)
print('全量对比：主表(偿付能力表)最低资本 缺失，但 明细表(最低资本表)有数')
print('=' * 80)
total_gap = 0
for seg, main_f, detail_f in PAIRS:
    print(f'\n### 板块：{seg} ###')
    main = read_main(main_f)
    if detail_f and os.path.exists(detail_f):
        detail = read_detail(detail_f)
    else:
        detail = {}
        print('  (无最低资本明细表，跳过明细对比)')
    # 找 主表缺 但 明细有
    gaps = []
    seen = set()
    for (comp, per), dval in detail.items():
        if dval is None or dval == '': continue
        mval = main.get((comp, per))
        if mval is None or mval == '':
            key = (comp, per)
            if key not in seen:
                seen.add(key)
                gaps.append((comp, per, dval, mval))
    gaps.sort()
    if gaps:
        print(f'  发现 {len(gaps)} 处「主表缺/空，明细有数」：')
        for comp, per, dval, mval in gaps:
            print(f'    - {comp} | {per} | 明细最低资本={dval}')
        total_gap += len(gaps)
    else:
        print('  无此类缺口')
print('\n' + '=' * 80)
print(f'总计：{total_gap} 处「主表最低资本缺，明细表有数」')
print('=' * 80)
