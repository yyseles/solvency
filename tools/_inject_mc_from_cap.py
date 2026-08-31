# -*- coding: utf-8 -*-
"""
从三个最低资本 Excel 提取 P/Q/R/S/T/U/V+N，写入 data.js 对应公司×期次。
覆盖规则：仅写入 P~V+N 字段，不动其他字段(C,D,E...I,J,K,L,M,O)。
数据源映射：
  产险最低资本.xlsx → segments.property
  人身险最低资本.xlsx → segments.life
  再保最低资本.xlsx   → segments.reins
字段映射：
  row 3  寿险业务保险风险最低资本合计       → P
  row 8  非寿险业务保险风险最低资本合计     → Q
  row 12 市场风险-最低资本合计              → R
  row 20 信用风险-最低资本合计              → S
  row 24 量化风险分散效应                   → T (负值)
  row 25 特定类别保险合同损失吸收效应        → U (负值)
  row 28 控制风险最低资本                   → V
  row 34 最低资本                           → N
"""

import json, re, sys, os
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_JS = os.path.join(BASE_DIR, '..', 'data.js')
EXCEL_DIR = os.path.join(BASE_DIR, '..', 'data')

SEG_FILES = {
    'property': os.path.join(EXCEL_DIR, '产险最低资本.xlsx'),
    'life':     os.path.join(EXCEL_DIR, '人身险最低资本.xlsx'),
    'reins':    os.path.join(EXCEL_DIR, '再保最低资本.xlsx'),
}

# 最低资本表行号映射 (0-indexed): 指标名 → 字段
MC_ROW_MAP = {
    3:  'P',   # 寿险业务保险风险最低资本合计
    8:  'Q',   # 非寿险业务保险风险最低资本合计
    12: 'R',   # 市场风险-最低资本合计
    20: 'S',   # 信用风险-最低资本合计
    24: 'T',   # 量化风险分散效应
    25: 'U',   # 特定类别保险合同损失吸收效应
    28: 'V',   # 控制风险最低资本
    34: 'N',   # 最低资本
}

def parse_excel_period(s):
    """'2022第1季度' → '2022Q1'"""
    m = re.match(r'(\d{4})第(\d)季度', str(s).strip())
    if m:
        return f"{m.group(1)}Q{m.group(2)}"
    return None

def load_min_cap_excel(filepath):
    """
    返回 { company_name: { period_key: { field: value } } }
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row = rows[0]  # row0: 指标名称 | 公司1 | 公司2 | ...
    period_row = rows[1]  # row1: 空 | 2022第1季度 | ...

    result = {}
    # 遍历每列：公司名跨列(空单元格=同前一家公司)，每列是一个期次
    current_comp = None
    for col_idx in range(1, len(header_row)):
        cell = header_row[col_idx]
        if cell and str(cell).strip():
            current_comp = str(cell).strip()
        if not current_comp:
            continue

        period_key = parse_excel_period(period_row[col_idx])
        if not period_key:
            continue

        values = {}
        for row_idx, field in MC_ROW_MAP.items():
            if row_idx < len(rows):
                cell_val = rows[row_idx][col_idx]
                if cell_val is not None and cell_val != '':
                    try:
                        values[field] = float(cell_val)
                    except (ValueError, TypeError):
                        pass

        if values:
            result.setdefault(current_comp, {})[period_key] = values

    return result


def main():
    print("=== 读取 data.js ===")
    with open(DATA_JS, 'r', encoding='utf-8') as f:
        txt = f.read()

    idx = txt.index('SOLVENCY_DATA = ') + len('SOLVENCY_DATA = ')
    depth = 0
    end = 0
    for i, c in enumerate(txt[idx:]):
        if c == '{':
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                end = i + 1
                break

    data = json.loads(txt[idx:idx + end])

    stats = {'updated': 0, 'skipped_no_period': 0, 'skipped_no_company': 0}

    for seg_key, filepath in SEG_FILES.items():
        if not os.path.exists(filepath):
            print(f"[WARN] 文件不存在: {filepath}, 跳过 {seg_key}")
            continue

        print(f"\n--- 处理 {seg_key}: {os.path.basename(filepath)} ---")
        excel_data = load_min_cap_excel(filepath)
        print(f"  Excel 中读取到 {len(excel_data)} 家公司")

        seg = data['segments'][seg_key]
        data_js_comps = set(seg['companies'])
        data_js_periods = {p['key'] for p in seg['periods']}

        matched = 0
        for comp_name, comp_periods in excel_data.items():
            if comp_name not in data_js_comps:
                stats['skipped_no_company'] += 1
                continue

            for pkey, values in comp_periods.items():
                target_key = pkey
                # Q4 数据若无法匹配，尝试映射到年度 key（如 2022Q4 → 2022）
                if pkey not in data_js_periods:
                    m = re.match(r'(\d{4})Q4$', pkey)
                    if m and m.group(1) in data_js_periods:
                        target_key = m.group(1)
                    else:
                        stats['skipped_no_period'] += 1
                        continue

                # 写入 P~V+N
                rec = seg['data'].setdefault(comp_name, {}).setdefault(target_key, {})
                for field, val in values.items():
                    rec[field] = val
                stats['updated'] += 1
                matched += 1

        print(f"  匹配写入: {matched} 条 (company×period)")
        unmatched_comps = set(excel_data.keys()) - data_js_comps
        if unmatched_comps:
            print(f"  Excel中有但data.js无此公司({len(unmatched_comps)}): {list(unmatched_comps)[:5]}")

    # 写回
    new_json = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
    new_txt = txt[:idx] + new_json + txt[idx + end:]
    with open(DATA_JS, 'w', encoding='utf-8') as f:
        f.write(new_txt)

    print(f"\n=== 完成 ===")
    print(f"  更新记录: {stats['updated']} 条")
    print(f"  跳过(期次不匹配): {stats['skipped_no_period']} 条")
    print(f"  跳过(公司不存在): {stats['skipped_no_company']} 条")
    print(f"  已写回: {DATA_JS}")


if __name__ == '__main__':
    main()
