# -*- coding: utf-8 -*-
"""中再集团本级偿付能力报告 PDF -> 结构化数据 -> 生成 0831 模板格式 xlsx（临时工具，不入 git）。
布局：主要指标页(万元) 取 C/D/E/F/G/H/I/J/K/L/M/N/O/V；九、最低资本明细页(元) 取 P/Q/R/S/T/U + mcDetail。
用法：
  python tools/_parse_zhongzai.py check [pdf_dir]      # 仅解析打印，供核对
  python tools/_parse_zhongzai.py gen [pdf_dir] [out]  # 生成 xlsx 到 out/data（默认 data/）
"""
import re, os, sys, glob
import pdfplumber
import openpyxl

REV = '0901'
PERIOD_STR = {
    '2022Q1': ('2022第一季度', '2022第1季度'), '2022Q2': ('2022第二季度', '2022第2季度'),
    '2022Q3': ('2022第三季度', '2022第3季度'), '2022Q4': ('2022第四季度', '2022第4季度'),
    '2023Q1': ('2023第一季度', '2023第1季度'),
    '2023Q2': ('2023第二季度', '2023第2季度'), '2023Q3': ('2023第三季度', '2023第3季度'),
    '2023Q4': ('2023第四季度', '2023第4季度'),
    '2024Q1': ('2024第一季度', '2024第1季度'), '2024Q2': ('2024第二季度', '2024第2季度'),
    '2024Q3': ('2024第三季度', '2024第3季度'), '2024Q4': ('2024第四季度', '2024第4季度'),
    '2025Q1': ('2025第一季度', '2025第1季度'), '2025Q2': ('2025第二季度', '2025第2季度'),
    '2025Q3': ('2025第三季度', '2025第3季度'), '2025Q4': ('2025第四季度', '2025第4季度'),
    '2026Q1': ('2026第一季度', '2026第1季度'),
}
# 偿付能力 xlsx row2 列标题 -> 字段字母
SOLV_COL_FROM_HEADER = {
    '综合偿付能力充足率': 'C', '核心偿付能力充足率': 'D', '综合偿付能力溢额': 'E', '核心偿付能力溢额': 'F',
    '认可资产': 'G', '认可负债': 'H', '实际资本': 'I', '核心一级资本': 'J', '核心二级资本': 'K',
    '附属一级资本': 'L', '附属二级资本': 'M', '最低资本': 'N', '量化风险最低资本': 'O',
    '寿险风险最低资本': 'P', '非寿险风险最低资本': 'Q', '市场风险最低资本': 'R', '信用风险最低资本': 'S',
    '量化风险分散效应': 'T', '特定类别保险合同损失吸收效应': 'U', '控制风险最低资本': 'V',
}
# 最低资本 xlsx 聚合标签 -> 字段
AGG_LABEL_TO_FIELD = {
    '量化风险最低资本': 'O', '寿险业务保险风险最低资本合计': 'P', '非寿险业务保险风险最低资本合计': 'Q',
    '市场风险-最低资本合计': 'R', '信用风险-最低资本合计': 'S', '量化风险分散效应': 'T',
    '特定类别保险合同损失吸收效应': 'U', '控制风险最低资本': 'V', '最低资本': 'N',
}
# 最低资本 xlsx 子项标签 -> mc key（与解析共用）
DETAIL_LABEL_TO_KEY = {
    '寿险业务保险风险-损失发生风险最低资本': 'mcP_loss', '寿险业务保险风险-退保风险最低资本': 'mcP_surr',
    '寿险业务保险风险-费用风险最低资本': 'mcP_exp', '寿险业务保险风险-风险分散效应': 'mcP_div',
    '非寿险业务保险风险-保费及准备金风险最低资本': 'mcQ_prem', '非寿险业务保险风险-巨灾风险最低资本': 'mcQ_cata',
    '非寿险业务保险风险-风险分散效应': 'mcQ_div', '市场风险-利率风险最低资本': 'mcR_rate',
    '市场风险-权益价格风险最低资本': 'mcR_eq', '市场风险-房地产价格风险最低资本': 'mcR_re',
    '市场风险-境外固定收益类资产价格风险最低资本': 'mcR_ofb', '市场风险-境外权益类资产价格风险最低资本': 'mcR_ofe',
    '市场风险-汇率风险最低资本': 'mcR_fx', '市场风险-风险分散效应': 'mcR_div',
    '信用风险-利差风险最低资本': 'mcS_spr', '信用风险-交易对手违约风险最低资本': 'mcS_def',
    '信用风险-风险分散效应': 'mcS_div', '附加资本': 'mcAdd', '逆周期附加资本': 'mcAdd_cyc',
    'D-SII附加资本': 'mcAdd_dsii', 'G-SII附加资本': 'mcAdd_gsii', '其他附加资本': 'mcAdd_oth',
}
# 老报告(2022)长标签被 pdfplumber 截断、数字转到下一行时的兜底核心子串（须唯一）
DETAIL_CORE = {
    '寿险业务保险风险-损失发生风险最低资本': '损失发生风险',
    '寿险业务保险风险-费用风险最低资本': '费用风险最低资本',
    '非寿险业务保险风险-保费及准备金风险最低资本': '保费及准备金',
    '非寿险业务保险风险-巨灾风险最低资本': '巨灾风险',
    '市场风险-境外固定收益类资产价格风险最低资本': '境外固定收益',
    '市场风险-境外权益类资产价格风险最低资本': '境外权益类资产价格风险',
    '市场风险-房地产价格风险最低资本': '房地产价格风险最低资本',
    '特定类别保险合同损失吸收效应': '特定类别保险合同损失吸收',
}

def _num(s):
    if s is None: return None
    s = str(s).replace(',', '').replace('%', '').strip()
    if s in ('-', ''): return 0.0
    try: return float(s)
    except: return None

def parse_pdf(path, prior=False):
    """解析中再集团本级偿付能力 PDF。

    prior=False -> 取「本季度数」列（第 1 个数字）
    prior=True  -> 取「上季度数/期初」列（第 2 个数字），用于生成**年度（审计后）**数据
                   （一季度报告的期初 = 上年四季度，且通常为年度审计后数）
    """
    pages = [p.extract_text() or '' for p in pdfplumber.open(path).pages]
    main = next((t for t in pages if '综合偿付能力充足率' in t and '（万元）' in t), '')
    IDX = 2 if prior else 1

    def pick(m):
        """按列取数：IDX=1 本季度 / IDX=2 期初。取不到该列则返回 None（不跨列兜底，避免取错列）。"""
        g = [x for x in (m.groups() if m else ()) if x is not None]
        return g[IDX - 1] if len(g) >= IDX else None

    NUM2 = r'\s*([\d,\.\-]+)(?:\s+([\d,\.\-]+))?'
    PCT2 = r'\s*([\d\.]+)%(?:\s+([\d\.]+)%)?'
    YUAN2 = r'(\d{1,3}(?:,\d{3})+(?:\.\d+)?)(?:\s+(\d{1,3}(?:,\d{3})+(?:\.\d+)?))?'
    PLAIN2 = r'(\d+(?:\.\d+)?)(?:\s+(\d+(?:\.\d+)?))?'
    rec = {}
    for f, lab in [('G', '认可资产'),
                   ('H', '认可负债'), ('I', '实际资本'), ('J', '核心一级资本'), ('K', '核心二级资本'),
                   ('L', '附属一级资本'), ('M', '附属二级资本'), ('N', '最低资本'),
                   ('O', '可资本化风险最低资本'), ('V', '控制风险最低资本')]:
        m = re.search(lab + r'（万元）' + NUM2, main)
        rec[f] = _num(pick(m)) if m else None
        if rec[f] is None:
            m2 = re.search(r'量化风险最低资本（万元）' + NUM2, main)
            rec[f] = _num(pick(m2)) if m2 else None
    # 溢额：直接取数（缺列再计算兜底）
    for f, lab in [('E', '综合偿付能力溢额'), ('F', '核心偿付能力溢额')]:
        m = re.search(lab + r'（万元）' + NUM2, main)
        rec[f] = _num(pick(m)) if m else None
    # 充足率：全文找标准 `xx偿付能力充足率（%）xx%`（2022Q2 第11页格式异常时会回退到第13页）
    all_text = '\n'.join(pages)
    for f, lab in [('C', '综合偿付能力充足率'), ('D', '核心偿付能力充足率')]:
        m = re.search(re.escape(lab) + r'（%）' + PCT2, all_text)
        v = pick(m)
        if v:
            rec[f] = round(float(v) / 100, 6)
    # 兜底：实际资本/最低资本 计算
    if rec.get('I') and rec.get('N'):
        if rec.get('C') is None:
            rec['C'] = round(rec['I'] / rec['N'], 6)
        if rec.get('D') is None:
            rec['D'] = round((rec.get('J', 0) + rec.get('K', 0)) / rec['N'], 6)
        if rec.get('E') is None:
            rec['E'] = round(rec['I'] - rec['N'], 2)
        if rec.get('F') is None:
            rec['F'] = round((rec.get('J', 0) + rec.get('K', 0)) - rec['N'], 2)
    detail_pages = [t for t in pages if '最低资本' in t and '风险最低资本' in t and '（万元）' not in t]
    merged = re.sub(r'\s+', ' ', '\n'.join(detail_pages))
    detail_lines = [ln for t in detail_pages for ln in t.splitlines()]

    def _r(s):
        return round(float(str(s).replace(',', '')) / 10000, 2)

    def get_yuan(label, core=None):
        # 1) 精确标签 + 千分位
        m = re.search(re.escape(label) + r'\D*?' + YUAN2, merged)
        v = pick(m)
        if v:
            return _r(v)
        # 2) 信用风险 标签变体
        alt = label.replace('信用风险-最低资本', '信用风险最低资本')
        if alt != label:
            m = re.search(re.escape(alt) + r'\D*?' + YUAN2, merged)
            v = pick(m)
            if v:
                return _r(v)
        # 3) 行级兜底（老报告长标签被截断，数字转到下一行）
        if core:
            for idx, line in enumerate(detail_lines):
                if core in line:
                    for j in range(idx, min(idx + 3, len(detail_lines))):
                        v = pick(re.search(YUAN2, detail_lines[j]))
                        if v:
                            return _r(v)
        # 4) 兜底：无千分位的纯数字（如 0）
        m = re.search(re.escape(label) + r'\D*?' + PLAIN2, merged)
        v = pick(m)
        if v:
            return _r(v)
        if core:
            for idx, line in enumerate(detail_lines):
                if core in line:
                    for j in range(idx, min(idx + 3, len(detail_lines))):
                        v = pick(re.search(PLAIN2, detail_lines[j]))
                        if v:
                            return _r(v)
        return None

    for f, lab in [('P', '寿险业务保险风险最低资本'), ('Q', '非寿险业务保险风险最低资本'),
                   ('R', '市场风险-最低资本'), ('S', '信用风险-最低资本'),
                   ('T', '量化风险分散效应'), ('U', '特定类别保险合同损失吸收效应')]:
        rec[f] = get_yuan(lab)
    mc = {k: get_yuan(lab, DETAIL_CORE.get(lab)) for lab, k in DETAIL_LABEL_TO_KEY.items()}
    return rec, mc

def generate(quarter, rec, mc, out):
    # 偿付能力 xlsx：复制 0831 模板，追加中再集团本级行
    wb = openpyxl.load_workbook('data/再保偿付能力26Q2-0831.xlsx')
    ws = wb.active
    hcol = {SOLV_COL_FROM_HEADER[v]: c for c in range(1, ws.max_column + 1)
            if (v := ws.cell(row=2, column=c).value) in SOLV_COL_FROM_HEADER}
    nr = ws.max_row + 1
    ws.cell(row=nr, column=1, value='中再集团本级')
    ws.cell(row=nr, column=2, value=PERIOD_STR[quarter][0])
    for f, col in hcol.items():
        ws.cell(row=nr, column=col, value=rec.get(f))
    wb.save(os.path.join(out, f'再保偿付能力{quarter}-{REV}.xlsx'))
    # 最低资本 xlsx：复制 0831 模板，新增中再集团本级列
    wb2 = openpyxl.load_workbook('data/再保最低资本26Q2-0831.xlsx')
    ws2 = wb2.active
    lrow = {ws2.cell(row=r, column=1).value: r for r in range(1, ws2.max_row + 1) if ws2.cell(row=r, column=1).value}
    nc = ws2.max_column + 1
    ws2.cell(row=1, column=nc, value='中再集团本级')
    ws2.cell(row=2, column=nc, value=PERIOD_STR[quarter][1])
    for lab, f in AGG_LABEL_TO_FIELD.items():
        if lab in lrow:
            ws2.cell(row=lrow[lab], column=nc, value=rec.get(f))
    for lab, k in DETAIL_LABEL_TO_KEY.items():
        if lab in lrow:
            ws2.cell(row=lrow[lab], column=nc, value=mc.get(k))
    wb2.save(os.path.join(out, f'再保最低资本{quarter}-{REV}.xlsx'))

def main():
    mode = sys.argv[1] if len(sys.argv) > 1 else 'check'
    d = sys.argv[2] if len(sys.argv) > 2 else 'data/_pdf_tmp'
    out = sys.argv[3] if len(sys.argv) > 3 else 'data'
    if mode == 'check':
        for path in sorted(glob.glob(os.path.join(d, '*.pdf'))):
            q = os.path.basename(path)[:-4]
            if q not in PERIOD_STR:
                continue
            rec, mc = parse_pdf(path)
            print('====', q, '====')
            print(' 主表:', {k: rec.get(k) for k in 'CDEFGHIJKLMNOPQRSTUV'})
            print(' mc:', mc)
    elif mode == 'gen':
        for path in sorted(glob.glob(os.path.join(d, '*.pdf'))):
            q = os.path.basename(path)[:-4]
            if q not in PERIOD_STR:
                continue
            rec, mc = parse_pdf(path)
            generate(q, rec, mc, out)
            print('generated', q)

if __name__ == '__main__':
    main()
